#!/usr/bin/env python3
"""
Excel Report Generator - Generate timing reports for JSON renaming operations
This module creates Excel reports with timing information for file renaming and Postman collection generation.

FEATURES:
- Track timing for naming convention operations
- Track timing for Postman collection generation
- Generate comprehensive Excel reports with timing data
- Support for multiple models and batch processing
- Calculate total time and average time per operation

REPORT COLUMNS:
- TC#ID: Test Case ID extracted from filename
- Model Name: Model name (WGS_CSBD, GBDF MCR, GBDF GRS)
- Edit ID: Edit identifier (e.g., rvn001, rvn002)
- EOB Code: EOB code (e.g., 00W5, 00W6)
- Naming Convention Time: Time taken for file renaming (ms)
- Postman Collection Time: Time taken for Postman collection generation (ms)
- Total Time: Sum of naming convention and Postman collection times (ms)
- Average Time: Average time per operation (ms)
"""

import os
import time
import json
from datetime import datetime
from typing import Dict, List, Any, Optional
from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows


class TimingTracker:
    """Track timing information for different operations."""
    
    def __init__(self):
        self.start_time = None
        self.end_time = None
        self.operation_name = ""
    
    def start(self, operation_name: str):
        """Start timing for an operation."""
        self.operation_name = operation_name
        self.start_time = time.time()
    
    def end(self) -> float:
        """End timing and return duration in milliseconds."""
        if self.start_time is None:
            return 0.0
        
        self.end_time = time.time()
        duration_ms = (self.end_time - self.start_time) * 1000
        return duration_ms
    
    def get_duration_ms(self) -> float:
        """Get current duration in milliseconds without ending the timer."""
        if self.start_time is None:
            return 0.0
        
        current_time = time.time()
        return (current_time - self.start_time) * 1000


class ExcelReportGenerator:
    """Generate Excel reports with timing information for JSON renaming operations."""
    
    def __init__(self, output_dir: str = "reports"):
        """
        Initialize the Excel report generator.
        
        Args:
            output_dir: Directory to save Excel reports
        """
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(exist_ok=True)
        
        # Initialize data storage
        self.timing_data = []
        self.current_session_data = []
        
        # Define column headers for the Excel report
        self.column_headers = [
            "TC#ID",
            "Model LOB", 
            "Model Name",
            "Edit ID",
            "EOB Code",
            "Naming Convention Time (ms)",
            "Postman Collection Time (ms)",
            "Total Time (ms)",
            "Average Time (ms)",
            "Timestamp",
            "Status"
        ]
    
    def start_timing_session(self, session_name: str = None):
        """Start a new timing session."""
        if session_name is None:
            session_name = f"Session_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        
        self.current_session_data = []
        print(f"[TIMING] Started timing session: {session_name}")
    
    def add_timing_record(self, 
                         tc_id: str,
                         model_lob: str,
                         model_name: str,
                         edit_id: str,
                         eob_code: str,
                         naming_convention_time_ms: float,
                         postman_collection_time_ms: float = 0.0,
                         status: str = "Success"):
        """
        Add a timing record to the current session.
        
        Args:
            tc_id: Test Case ID
            model_lob: Model LOB (WGS_CSBD, GBDF_MCR, GBDF_GRS)
            model_name: Model name (Covid, Multiple E&M Same day, etc.)
            edit_id: Edit identifier
            eob_code: EOB code
            naming_convention_time_ms: Time taken for naming convention (ms)
            postman_collection_time_ms: Time taken for Postman collection (ms)
            status: Operation status (Success, Failed, etc.)
        """
        total_time = naming_convention_time_ms + postman_collection_time_ms
        average_time = total_time / 2 if total_time > 0 else 0.0
        
        record = {
            "TC#ID": tc_id,
            "Model LOB": model_lob,
            "Model Name": model_name,
            "Edit ID": edit_id,
            "EOB Code": eob_code,
            "Naming Convention Time (ms)": round(naming_convention_time_ms, 2),
            "Postman Collection Time (ms)": round(postman_collection_time_ms, 2),
            "Total Time (ms)": round(total_time, 2),
            "Average Time (ms)": round(average_time, 2),
            "Timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "Status": status
        }
        
        self.current_session_data.append(record)
        self.timing_data.append(record)
        
        print(f"[TIMING] Added record: {tc_id} - {model_lob} - {model_name} - Total: {total_time:.2f}ms")
    
    def generate_excel_report(self, filename: str = None, model_type: str = None) -> str:
        """
        Generate Excel report with timing data.
        
        Args:
            filename: Custom filename for the report
            model_type: Type of model (WGS_CSBD, GBDF_MCR, GBDF_GRS) for filename generation
            
        Returns:
            Path to the generated Excel file
        """
        if not self.timing_data:
            print("[WARNING] No timing data available for report generation")
            return None
        
        # Generate filename if not provided
        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            if model_type:
                filename = f"JSON_Renaming_Timing_Report_{model_type}_{timestamp}.xlsx"
            else:
                filename = f"JSON_Renaming_Timing_Report_{timestamp}.xlsx"
        
        # Ensure filename has .xlsx extension
        if not filename.endswith('.xlsx'):
            filename += '.xlsx'
        
        report_path = self.output_dir / filename
        
        try:
            # Create DataFrame from timing data
            df = pd.DataFrame(self.timing_data)
            
            # Reorder columns to match the defined headers
            df = df.reindex(columns=self.column_headers)
            
            # Create Excel workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Timing Report"
            
            # Add data to worksheet
            for r in dataframe_to_rows(df, index=False, header=True):
                ws.append(r)
            
            # Apply formatting
            self._apply_excel_formatting(ws, len(self.timing_data))
            
            # Add summary statistics
            self._add_summary_sheet(wb, df)
            
            # Save the workbook
            wb.save(report_path)
            
            print(f"[SUCCESS] Excel report generated: {report_path}")
            print(f"[INFO] Total records: {len(self.timing_data)}")
            
            return str(report_path)
            
        except Exception as e:
            print(f"[ERROR] Failed to generate Excel report: {e}")
            return None
    
    def _apply_excel_formatting(self, ws, num_rows: int):
        """Apply formatting to the Excel worksheet."""
        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_alignment = Alignment(horizontal='center', vertical='center')
        
        # Format header row
        for col in range(1, len(self.column_headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = center_alignment
        
        # Format data rows
        for row in range(2, num_rows + 2):
            for col in range(1, len(self.column_headers) + 1):
                cell = ws.cell(row=row, column=col)
                cell.border = border
                if col in [5, 6, 7, 8]:  # Time columns
                    cell.alignment = center_alignment
                    cell.number_format = '0.00'
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _add_summary_sheet(self, wb, df):
        """Add summary statistics sheet."""
        ws_summary = wb.create_sheet("Summary Statistics")
        
        # Calculate summary statistics
        total_records = len(df)
        total_naming_time = df['Naming Convention Time (ms)'].sum()
        total_postman_time = df['Postman Collection Time (ms)'].sum()
        total_time = df['Total Time (ms)'].sum()
        avg_naming_time = df['Naming Convention Time (ms)'].mean()
        avg_postman_time = df['Postman Collection Time (ms)'].mean()
        avg_total_time = df['Total Time (ms)'].mean()
        
        # Model breakdown
        model_lob_counts = df['Model LOB'].value_counts()
        model_name_counts = df['Model Name'].value_counts()
        
        # Status breakdown
        status_counts = df['Status'].value_counts()
        
        # Add summary data
        summary_data = [
             ["SUMMARY STATISTICS", ""],
             ["", ""],
             ["Total Records", total_records],
             ["", ""],
             ["TIMING STATISTICS", ""],
             ["Total Naming Convention Time (ms)", f"{total_naming_time:.2f}"],
             ["Total Postman Collection Time (ms)", f"{total_postman_time:.2f}"],
             ["Total Time (ms)", f"{total_time:.2f}"],
             ["Average Naming Convention Time (ms)", f"{avg_naming_time:.2f}"],
             ["Average Postman Collection Time (ms)", f"{avg_postman_time:.2f}"],
             ["Average Total Time (ms)", f"{avg_total_time:.2f}"],
             ["", ""],
             ["MODEL LOB BREAKDOWN", ""],
         ]
         
        for model_lob, count in model_lob_counts.items():
            summary_data.append([f"{model_lob} Records", count])
        
        summary_data.extend([
            ["", ""],
            ["MODEL NAME BREAKDOWN", ""],
        ])
        
        for model_name, count in model_name_counts.items():
            summary_data.append([f"{model_name} Records", count])
        
        summary_data.extend([
            ["", ""],
            ["STATUS BREAKDOWN", ""],
        ])
        
        for status, count in status_counts.items():
            summary_data.append([f"{status} Records", count])
        
        # Add data to summary sheet
        for row_data in summary_data:
            ws_summary.append(row_data)
        
        # Format summary sheet
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        for row in range(1, len(summary_data) + 1):
            cell = ws_summary.cell(row=row, column=1)
            if cell.value and "STATISTICS" in str(cell.value) or "BREAKDOWN" in str(cell.value):
                cell.font = header_font
                cell.fill = header_fill
        
        # Auto-adjust column widths
        ws_summary.column_dimensions['A'].width = 40
        ws_summary.column_dimensions['B'].width = 20
    
    def get_session_summary(self) -> Dict[str, Any]:
        """Get summary of current session data."""
        if not self.current_session_data:
            return {"message": "No data in current session"}
        
        df = pd.DataFrame(self.current_session_data)
        
        return {
            "total_records": len(self.current_session_data),
            "total_naming_time_ms": df['Naming Convention Time (ms)'].sum(),
            "total_postman_time_ms": df['Postman Collection Time (ms)'].sum(),
            "total_time_ms": df['Total Time (ms)'].sum(),
            "average_time_ms": df['Total Time (ms)'].mean(),
            "model_lobs": df['Model LOB'].unique().tolist(),
            "model_names": df['Model Name'].unique().tolist(),
            "status_counts": df['Status'].value_counts().to_dict()
        }
    
    def clear_data(self):
        """Clear all timing data."""
        self.timing_data = []
        self.current_session_data = []
        print("[INFO] Timing data cleared")
    
    def export_to_csv(self, filename: str = None) -> str:
        """
        Export timing data to CSV format.
        
        Args:
            filename: Custom filename for the CSV file
            
        Returns:
            Path to the generated CSV file
        """
        if not self.timing_data:
            print("[WARNING] No timing data available for CSV export")
            return None
        
        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"JSON_Renaming_Timing_Report_{timestamp}.csv"
        
        if not filename.endswith('.csv'):
            filename += '.csv'
        
        csv_path = self.output_dir / filename
        
        try:
            df = pd.DataFrame(self.timing_data)
            df.to_csv(csv_path, index=False)
            
            print(f"[SUCCESS] CSV report generated: {csv_path}")
            return str(csv_path)
            
        except Exception as e:
            print(f"[ERROR] Failed to generate CSV report: {e}")
            return None


# Global instance for easy access
excel_reporter = ExcelReportGenerator()


def create_timing_tracker() -> TimingTracker:
    """Create a new timing tracker instance."""
    return TimingTracker()


def get_excel_reporter() -> ExcelReportGenerator:
    """Get the global Excel reporter instance."""
    return excel_reporter

def create_excel_reporter_for_model_type(model_type: str) -> ExcelReportGenerator:
    """
    Create a new Excel reporter instance for a specific model type.
    This ensures separate data collection for different model types.
    
    Args:
        model_type: Type of model (WGS_CSBD, GBDF_MCR, GBDF_GRS)
        
    Returns:
        New ExcelReportGenerator instance
    """
    return ExcelReportGenerator()


if __name__ == "__main__":
    """
    Test the Excel report generator functionality.
    """
    # Create test data
    reporter = ExcelReportGenerator()
    reporter.start_timing_session("Test Session")
    
    # Add some test records
    reporter.add_timing_record(
        tc_id="TS_01_12345",
        model_lob="WGS_CSBD",
        model_name="Covid",
        edit_id="rvn001",
        eob_code="W04",
        naming_convention_time_ms=150.5,
        postman_collection_time_ms=75.2
    )
    
    reporter.add_timing_record(
        tc_id="TS_47_99202",
        model_lob="GBDF_MCR",
        model_name="Covid",
        edit_id="rvn001",
        eob_code="v04",
        naming_convention_time_ms=200.3,
        postman_collection_time_ms=100.1
    )
    
    # Generate reports
    excel_path = reporter.generate_excel_report()
    csv_path = reporter.export_to_csv()
    
    # Print summary
    summary = reporter.get_session_summary()
    print(f"\nSession Summary: {summary}")
    
    print(f"\nTest completed successfully!")
    print(f"Excel report: {excel_path}")
    print(f"CSV report: {csv_path}")
